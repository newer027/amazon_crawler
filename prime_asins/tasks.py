from openpyxl import load_workbook
from openpyxl.compat import range
from celery import Celery
import time
from prime_asins.models import Asin_detail
from celery.task.schedules import crontab
from celery.decorators import periodic_task
from prime_asins.asin_detail import asin_to_detail
from celery.utils.log import get_task_logger


app = Celery()
app.config_from_object('django.conf:settings')

@app.task(bind=True,time_limit=4000,acks_late=True,default_retry_delay=30,max_retries=3)
def import_user(self, filename):
    wb = load_workbook(filename=filename)
    ws = wb.get_sheet_names()
    country = ''
    if ws[0][0] == '美':
        country = 'us'
    elif ws[0][0] == '加':
        country = 'ca'
    elif ws[0][0] == '法':
        country = 'fr'
    elif ws[0][0] == '德':
        country = 'de'
    elif ws[0][0] == '意':
        country = 'it'
    elif ws[0][0] == '日':
        country = 'jp'
    elif ws[0][0] == '西':
        country = 'es'
    elif ws[0][0] == '英':
        country = 'uk'
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['description', 'sell_rank', 'asin', 'title', 'fu_asin', 'fu_asin_title',
               'sell_quant', 'sell_amount','fu_sell_quant','fu_sell_amout']
    for row in range(3, ws.max_row):
        cell = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            cell[key] = ws.cell(row=row, column=col).value
        if not Asin_detail.objects.filter(description = cell['description'],country=country,fu_asin_title = cell['fu_asin_title']):
            asin_detail = Asin_detail(description = cell['description'], sell_rank = cell['sell_rank'],
                              asin = cell['asin'], title = cell['title'], fu_asin = cell['fu_asin'],
                              fu_asin_title = cell['fu_asin_title'], sell_quant = cell['sell_quant'],
                              sell_amount = cell['sell_amount'], fu_sell_quant = cell['fu_sell_quant'],
                              fu_sell_amout = cell['fu_sell_amout'],country=country)
            asin_detail.save()
        if not row % 1000:
            print("row:"+str(row))


@app.task(bind=True,time_limit=50000,acks_late=True,default_retry_delay=30,max_retries=3)
def fu_asin_rank(self):
    count = 0
    for asin in Asin_detail.objects.all():
        count += 1
        index = 0
        for refer_asin in Asin_detail.objects.filter(country=asin.country,description=asin.description).order_by('-fu_sell_amout'):
            index += 1
            if asin.fu_asin == refer_asin.fu_asin:
                asin.sell_rank=index
                asin.save()
                break
        if not count % 200:
            print("count:"+str(count))


def fu_asin_rank_0803():
    countries = [country[0] for country in Asin_detail.COUNTRY_CHOICES]
    for country in countries:
        categories = [category[0] for category in Asin_detail.objects.filter(country=country).values_list('description').distinct('description')]
        for category in categories:
            print("国家:",country,"分类:",category)
            asin_details=Asin_detail.objects.filter(country=country, description=category).order_by('-fu_sell_amout')
            for index in range(len(asin_details)):
                asin_details[index].sell_rank=index+1
                asin_details[index].save()


"""
logger = get_task_logger(__name__)
@periodic_task(
    run_every=(crontab(minute=10,hour='*/3')),
    name="prime_asins_depart",
    ignore_result=True
)
def prime_asins_depart():
    start = time.time()
    for asin_detail in Asin_detail.objects.filter(img_url=None)[0:2500]:
        if not asin_detail.img_url and (time.time()-start)<60*60*0.7:
            asin_to_detail.delay(asin_detail.fu_asin,asin_detail.country)
            time.sleep(1.5)
    logger.info("Saved latest prime_asins_depart")
"""